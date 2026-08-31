from __future__ import annotations

import csv
from pathlib import Path


def load_rows(path: str | Path, *, sheet_index: int | None = None) -> list[list[str]]:
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _load_xlsx(path, sheet_index if sheet_index is not None else 0)
    if suffix in {".csv", ".tsv", ".txt"}:
        delim = "\t" if suffix == ".tsv" else ","
        return _load_csv(path, delimiter=delim)
    raise ValueError(f"Неподдерживаемый формат: {suffix or 'без расширения'}")


def load_xlsx_sheet_count(path: str | Path) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        return len(wb.worksheets)
    finally:
        wb.close()


def _load_xlsx(path: Path, sheet_index: int) -> list[list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_index < 0 or sheet_index >= len(wb.worksheets):
            raise IndexError(f"Лист {sheet_index} отсутствует")
        ws = wb.worksheets[sheet_index]
        return [
            ["" if v is None else str(v) for v in row]
            for row in ws.iter_rows(values_only=True)
        ]
    finally:
        wb.close()


def _load_csv(path: Path, *, delimiter: str = ",") -> list[list[str]]:
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            with path.open(newline="", encoding=enc) as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
                    delimiter = dialect.delimiter
                except csv.Error:
                    pass
                return list(csv.reader(f, delimiter=delimiter))
        except UnicodeDecodeError:
            continue
    raise ValueError("Не удалось прочитать CSV (кодировка)")
